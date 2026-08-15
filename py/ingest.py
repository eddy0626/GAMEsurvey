# -*- coding: utf-8 -*-
"""응답 시트 읽기 · 헤더 매핑 · 응답 파싱

src/Aggregate.gs 의 [A] 순수 계산 함수 중 파싱 부분을 그대로 옮긴 것.
구글 폼이 문항 제목에 붙이는 연속번호를 떼고 본문으로만 매칭한다.
게임마다 고유 문항 개수가 10/10/13/7/17 로 달라 같은 공통 문항이라도
번호가 다르기 때문이다. 열 위치는 어디에서도 하드코딩하지 않는다.
"""
import csv
import re
import unicodedata

from . import config as C

_번호 = re.compile(r"^\s*\d+(?:[-.]\d+)*\s*[.)]\s*")
_따옴표1 = str.maketrans("‘’‛ʼ", "''''")
_따옴표2 = str.maketrans("“”‟", '"""')
_대시 = str.maketrans("-‐‑‒–—―−－", "—" * 9)
_중점 = str.maketrans("·•・･", "····")
_공백 = str.maketrans(" 　​", "   ")


def 정규화(s):
    """헤더 · 문항 텍스트 비교용 정규화"""
    if s is None:
        return ""
    t = unicodedata.normalize("NFC", str(s))
    t = _번호.sub("", t)
    t = t.translate(_따옴표1).translate(_따옴표2).translate(_대시).translate(_중점).translate(_공백)
    return re.sub(r"\s+", " ", t).strip()


def 헤더맵(헤더행):
    """헤더 행 → (공통 {키: 열}, 고유 [(문항, 열)], 누락 [키], 중복 [키])"""
    정규 = [정규화(h) for h in 헤더행]
    역 = {}
    for 키, 후보들 in C.공통문항.items():
        for 후보 in 후보들:
            역[정규화(후보)] = 키

    공통, 사용, 중복 = {}, set(), []
    for i, t in enumerate(정규):
        키 = 역.get(t)
        if 키 is None:
            continue
        if 키 in 공통:
            중복.append(키)
            continue
        공통[키] = i
        사용.add(i)

    고유 = [(정규[i], i) for i in range(len(정규)) if i not in 사용 and 정규[i]]
    누락 = [k for k in C.공통문항 if k not in 공통]
    return 공통, 고유, 누락, 중복


def 문자(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(v)).translate(_공백)).strip()


def 숫자(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return v
    s = re.sub(r"[^\d.\-]", "", str(v))
    if s in ("", "-", "."):
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f == int(f) else f


def 복수답분해(v):
    """구글 폼은 체크박스 답을 ', ' 로 이어 붙인다.
    '10,000~14,900원' 처럼 선택지 안에 쉼표가 있는 객관식을 잘못 쪼개지 않도록
    쉼표 뒤 공백이 있을 때만 나눈다."""
    s = 문자(v)
    if not s:
        return []
    return [p.strip() for p in re.split(r",\s+", s) if p.strip()]


_무응답 = re.compile(r"^(없음|없습니다|없어요|없다|무|-|\.|x|X|N/A|na)$", re.I)


def 무응답(s):
    t = 문자(s)
    return (not t) or bool(_무응답.match(re.sub(r"\s", "", t)))


class 응답:
    """응답 한 건"""
    __slots__ = ("행", "게임코드", "게임명", "타임스탬프", "이름", "ID", "유형",
                 "연령대", "성별", "선호장르", "육각", "진행흐름", "기술안정성",
                 "진행도", "개선점", "구매의향", "적정가격", "추천의향",
                 "버그경험", "버그상황", "개발팀메시지", "고유")

    def __getitem__(self, k):
        return getattr(self, k)


def 파싱(헤더행, 데이터행, 게임):
    """응답 시트 → [응답]"""
    공통, 고유정의, 누락, 중복 = 헤더맵(헤더행)
    코드 = 게임["코드"]
    체크셋 = {정규화(t) for t in C.체크박스문항.get(코드, [])}

    def 셀(행, 키):
        i = 공통.get(키)
        return "" if i is None or i >= len(행) else 행[i]

    나온것 = []
    for n, 행 in enumerate(데이터행):
        if not any(문자(v) for v in 행):
            continue
        r = 응답()
        r.행 = n + 2
        r.게임코드, r.게임명 = 코드, 게임["게임명"]
        r.타임스탬프 = 셀(행, "타임스탬프")
        r.이름 = 문자(셀(행, "이름"))
        r.ID, r.유형 = "", ""
        r.연령대 = 문자(셀(행, "연령대"))
        r.성별 = 문자(셀(행, "성별"))
        r.선호장르 = 복수답분해(셀(행, "선호장르"))
        r.육각 = [숫자(셀(행, k)) for k, _ in C.육각축]
        r.진행흐름 = 숫자(셀(행, "진행흐름"))
        r.기술안정성 = 숫자(셀(행, "기술안정성"))
        r.진행도 = 문자(셀(행, "진행도"))
        r.개선점 = 문자(셀(행, "개선점"))
        r.구매의향 = 문자(셀(행, "구매의향"))
        r.적정가격 = 문자(셀(행, "적정가격"))
        r.추천의향 = 숫자(셀(행, "추천의향"))
        r.버그경험 = 문자(셀(행, "버그경험"))
        r.버그상황 = 문자(셀(행, "버그상황"))
        r.개발팀메시지 = 문자(셀(행, "개발팀메시지"))

        r.고유 = []
        for 문항, 열 in 고유정의:
            원값 = 문자(행[열]) if 열 < len(행) else ""
            체크 = 문항 in 체크셋
            r.고유.append(dict(문항=문항, 열=열, 체크박스=체크, 답=원값,
                               값=복수답분해(원값) if 체크 else ([원값] if 원값 else [])))
        나온것.append(r)

    return 나온것, dict(공통=공통, 고유=고유정의, 누락=누락, 중복=중복)


# ── 파일에서 읽기 ────────────────────────────────────────────────────────

def CSV읽기(경로):
    with open(경로, encoding="utf-8-sig", newline="") as fp:
        rows = list(csv.reader(fp))
    return (rows[0], rows[1:]) if rows else ([], [])


def XLSX읽기(경로, 탭=None):
    from openpyxl import load_workbook
    wb = load_workbook(경로, data_only=True)
    ws = wb[탭] if 탭 else wb.worksheets[0]
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    return (list(rows[0]), [list(r) for r in rows[1:]]) if rows else ([], [])


def 자동읽기(경로, 탭=None):
    return XLSX읽기(경로, 탭) if str(경로).lower().endswith((".xlsx", ".xlsm")) else CSV읽기(경로)


# ── 명부 ─────────────────────────────────────────────────────────────────

def 이름키(s):
    return re.sub(r"\s+", "", 문자(s))


def 명부적용(응답목록, 명부):
    """명부(이름 → dict(ID, 유형))를 응답에 붙인다.
    명부에 없는 이름은 다음 ID 를 발급하고 유형을 '미상' 으로 둔다."""
    쓰인ID = {v["ID"] for v in 명부.values() if v.get("ID")}

    def 다음ID():
        n = 1
        while True:
            cand = C.응답자["접두사"] + str(n).zfill(C.응답자["자리수"])
            if cand not in 쓰인ID:
                쓰인ID.add(cand)
                return cand
            n += 1

    새로등록 = []
    for r in 응답목록:
        키 = 이름키(r.이름)
        if not 키:
            r.ID, r.유형 = "", C.응답자["미상유형"]
            continue
        if 키 not in 명부:
            명부[키] = dict(이름=r.이름, ID=다음ID(), 유형=C.응답자["미상유형"])
            새로등록.append(명부[키])
        elif not 명부[키].get("ID"):
            명부[키]["ID"] = 다음ID()
        r.ID = 명부[키]["ID"]
        r.유형 = 명부[키].get("유형") or C.응답자["기본유형"]
    return 새로등록
